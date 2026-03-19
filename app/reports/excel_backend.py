from __future__ import annotations

from collections import Counter, defaultdict
from collections.abc import Mapping
from dataclasses import dataclass
from datetime import date, datetime, time, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from app.schemas.reports import ReportFilters, ReportMetric, ReportRequest, ReportResult, ReportType
from app.schemas.tools import RunReportResponse

COLUMN_MAP: dict[str, str] = {
    "Կտրոն": "check",
    "Սեղաններ / Սրահներ": "source",
    "Սրահներ / Սեղաններ": "source",
    "Ամսաթիվ": "order_date",
    "Աոաքիչ": "courier",
    "Առաքիչ": "courier",
    "Մատուցող": "waiter",
    "Հաճախորդ": "customer",
    "Հասցե": "address",
    "Հեռախոսահամար": "phone_number",
    "Ինքնարժեք": "net_cost",
    "Հաշիվ": "invoice",
    "Վճարված": "paid",
    "Առաքման գումար": "delivery_fee",
    "Գումար": "total_amount",
    "Նկարագրություն": "description",
}

_REQUIRED_COLUMNS = {"order_date", "total_amount"}
EXCEL_BACKEND_WARNING = "excel_backend_file_data"


@dataclass(frozen=True)
class ExcelOrderRow:
    order_date: date
    total_amount: Decimal
    check: str | None = None
    source: str | None = None
    courier: str | None = None
    waiter: str | None = None
    customer: str | None = None
    address: str | None = None
    phone_number: str | None = None
    net_cost: Decimal | None = None
    invoice: Decimal | None = None
    paid_amount: Decimal | None = None
    paid: bool | None = None
    delivery_fee: Decimal | None = None
    description: str | None = None


def _normalize_header(value: Any) -> str:
    text = str(value or "").replace("\u00A0", " ").strip()
    return " ".join(text.split()).lower()


def _to_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _extract_phone_digits(value: str | None) -> str | None:
    if value is None:
        return None
    digits = "".join(ch for ch in value if ch.isdigit())
    if len(digits) < 6 or len(digits) > 15:
        return None
    if digits.startswith("374") and len(digits) in {11, 12}:
        return f"0{digits[3:]}"
    return digits


def _extract_name_text(value: str | None) -> str | None:
    if value is None:
        return None
    text = " ".join(value.split()).strip()
    if not text:
        return None
    letters = sum(ch.isalpha() for ch in text)
    if letters == 0:
        return None
    digits = sum(ch.isdigit() for ch in text)
    if digits > letters:
        return None
    return text


def _parse_contact_fields(
    *,
    raw_customer: str | None,
    raw_phone: str | None,
) -> tuple[str | None, str | None]:
    phone_number = _extract_phone_digits(raw_phone) or _extract_phone_digits(raw_customer)
    customer_name = _extract_name_text(raw_customer)
    if customer_name is None:
        customer_name = _extract_name_text(raw_phone)
    return customer_name, phone_number


def _parse_date(value: Any, *, row_number: int, field_name: str) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value)
        except Exception as exc:
            raise ValueError(f"Row {row_number}: invalid {field_name} value '{value}'.") from exc
        if isinstance(converted, datetime):
            return converted.date()
        if isinstance(converted, date):
            return converted
    if isinstance(value, str):
        raw = value.strip()
        formats = (
            "%d.%m.%y %H:%M",
            "%d.%m.%Y %H:%M",
            "%d.%m.%y",
            "%d.%m.%Y",
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y-%m-%d",
        )
        for fmt in formats:
            try:
                return datetime.strptime(raw, fmt).date()
            except ValueError:
                continue
    raise ValueError(f"Row {row_number}: invalid {field_name} value '{value}'.")


def _parse_decimal(value: Any, *, row_number: int, field_name: str) -> Decimal:
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            raise ValueError(f"Row {row_number}: empty {field_name} value.")
        raw = raw.replace("֏", "").replace("$", "").replace("€", "")
        raw = raw.replace("\u00A0", "").replace(" ", "")
        if "," in raw and "." not in raw:
            raw = raw.replace(",", ".")
        elif "," in raw and "." in raw:
            if raw.rfind(",") > raw.rfind("."):
                raw = raw.replace(".", "").replace(",", ".")
            else:
                raw = raw.replace(",", "")
        try:
            return Decimal(raw)
        except InvalidOperation as exc:
            raise ValueError(
                f"Row {row_number}: invalid {field_name} value '{value}'."
            ) from exc
    raise ValueError(f"Row {row_number}: invalid {field_name} value '{value}'.")


def _parse_optional_decimal(
    value: Any,
    *,
    row_number: int,
    field_name: str,
) -> Decimal | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    return _parse_decimal(value, row_number=row_number, field_name=field_name)


def _parse_optional_bool(value: Any) -> bool | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if not text:
        return None
    if text in {"1", "true", "yes", "y", "paid", "վճարված", "այո"}:
        return True
    if text in {"0", "false", "no", "n", "unpaid", "չվճարված", "ոչ"}:
        return False
    try:
        return Decimal(text) > 0
    except InvalidOperation:
        return None


def load_excel_orders(
    file_path: str | Path,
    *,
    sheet_name: str | None = None,
    column_map: Mapping[str, str] = COLUMN_MAP,
) -> list[ExcelOrderRow]:
    path = Path(file_path)
    workbook = load_workbook(path, data_only=True, read_only=True)

    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook.active

        header_row_idx: int | None = None
        header_values: list[Any] | None = None
        for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            if any(cell is not None and str(cell).strip() for cell in row):
                header_row_idx = row_idx
                header_values = list(row)
                break

        if header_row_idx is None or header_values is None:
            raise ValueError("Excel file has no header row.")

        normalized_map = {
            _normalize_header(non_english): english
            for non_english, english in column_map.items()
        }

        target_to_index: dict[str, int] = {}
        for idx, header_cell in enumerate(header_values):
            key = _normalize_header(header_cell)
            target = normalized_map.get(key)
            if target:
                target_to_index[target] = idx

        missing = sorted(col for col in _REQUIRED_COLUMNS if col not in target_to_index)
        if missing:
            raise ValueError(f"Missing required mapped columns: {', '.join(missing)}")

        def get_cell(row: tuple[Any, ...], field: str) -> Any:
            index = target_to_index.get(field)
            if index is None or index >= len(row):
                return None
            return row[index]

        output: list[ExcelOrderRow] = []
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=header_row_idx + 1, values_only=True),
            start=header_row_idx + 1,
        ):
            if all(
                cell is None or (isinstance(cell, str) and not cell.strip())
                for cell in row
            ):
                continue

            raw_customer = _to_text(get_cell(row, "customer"))
            raw_phone = _to_text(get_cell(row, "phone_number"))
            customer_name, phone_number = _parse_contact_fields(
                raw_customer=raw_customer,
                raw_phone=raw_phone,
            )

            output.append(
                ExcelOrderRow(
                    order_date=_parse_date(
                        get_cell(row, "order_date"),
                        row_number=row_number,
                        field_name="order_date",
                    ),
                    total_amount=_parse_decimal(
                        get_cell(row, "total_amount"),
                        row_number=row_number,
                        field_name="total_amount",
                    ),
                    check=_to_text(get_cell(row, "check")),
                    source=_to_text(get_cell(row, "source")),
                    courier=_to_text(get_cell(row, "courier")),
                    waiter=_to_text(get_cell(row, "waiter")),
                    customer=customer_name,
                    address=_to_text(get_cell(row, "address")),
                    phone_number=phone_number,
                    net_cost=_parse_optional_decimal(
                        get_cell(row, "net_cost"),
                        row_number=row_number,
                        field_name="net_cost",
                    ),
                    invoice=_parse_optional_decimal(
                        get_cell(row, "invoice"),
                        row_number=row_number,
                        field_name="invoice",
                    ),
                    paid_amount=_parse_optional_decimal(
                        get_cell(row, "paid"),
                        row_number=row_number,
                        field_name="paid",
                    ),
                    paid=_parse_optional_bool(get_cell(row, "paid")),
                    delivery_fee=_parse_optional_decimal(
                        get_cell(row, "delivery_fee"),
                        row_number=row_number,
                        field_name="delivery_fee",
                    ),
                    description=_to_text(get_cell(row, "description")),
                )
            )

        return output
    finally:
        workbook.close()


def _generated_at(filters: ReportFilters) -> datetime:
    return datetime.combine(filters.date_to, time.min, tzinfo=timezone.utc)


def _rows_in_date_range(rows: list[ExcelOrderRow], filters: ReportFilters) -> list[ExcelOrderRow]:
    return [row for row in rows if filters.date_from <= row.order_date <= filters.date_to]


def _normalize_key(value: str | None) -> str | None:
    if value is None:
        return None
    normalized = " ".join(value.split()).strip().lower()
    return normalized or None


def _customer_identity_key(row: ExcelOrderRow) -> str | None:
    return row.phone_number


def _group_sales_by_customer(rows: list[ExcelOrderRow]) -> dict[str, Decimal]:
    grouped: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    for row in rows:
        customer_key = _customer_identity_key(row)
        if customer_key is None:
            continue
        grouped[customer_key] += _sales_amount(row)
    return grouped


def _sales_amount(row: ExcelOrderRow) -> Decimal:
    # In real exported files, invoice carries monetary sales while total_amount is often zero.
    if row.invoice is not None:
        return row.invoice
    return row.total_amount


def _sales_total(rows: list[ExcelOrderRow]) -> Decimal:
    return sum((_sales_amount(row) for row in rows), start=Decimal("0"))


def _order_count(rows: list[ExcelOrderRow]) -> int:
    checks = {row.check for row in rows if row.check}
    if checks:
        return len(checks)
    return len(rows)


def _group_sales(
    rows: list[ExcelOrderRow],
    *,
    key_getter: Any,
) -> dict[str, Decimal]:
    grouped: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    for row in rows:
        raw = key_getter(row)
        label = " ".join(raw.split()).strip() if isinstance(raw, str) else ""
        label = label or "unknown"
        grouped[label] += _sales_amount(row)
    return grouped


def _ranked_metrics(grouped: dict[str, Decimal], *, top_n: int | None = None) -> list[ReportMetric]:
    ranked = sorted(
        grouped.items(),
        key=lambda item: (-item[1], item[0].lower()),
    )
    if top_n is not None:
        ranked = ranked[:top_n]
    return [ReportMetric(label=label, value=float(value)) for label, value in ranked]


def _daily_sales_metrics(rows: list[ExcelOrderRow]) -> list[ReportMetric]:
    totals: dict[date, Decimal] = defaultdict(lambda: Decimal("0"))
    for row in rows:
        totals[row.order_date] += _sales_amount(row)
    return [
        ReportMetric(label=day.isoformat(), value=float(total))
        for day, total in sorted(totals.items())
    ]


def _daily_order_metrics(rows: list[ExcelOrderRow]) -> list[ReportMetric]:
    counts: dict[date, int] = defaultdict(int)
    for row in rows:
        counts[row.order_date] += 1
    return [
        ReportMetric(label=day.isoformat(), value=float(count))
        for day, count in sorted(counts.items())
    ]


def _sales_by_weekday_metrics(rows: list[ExcelOrderRow]) -> list[ReportMetric]:
    weekday_names = ("monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday")
    totals: dict[int, Decimal] = defaultdict(lambda: Decimal("0"))
    for row in rows:
        totals[row.order_date.weekday()] += _sales_amount(row)
    return [
        ReportMetric(label=weekday_names[index], value=float(totals[index]))
        for index in range(7)
    ]


def _metrics_for_request(
    request: ReportRequest,
    rows: list[ExcelOrderRow],
    *,
    all_rows: list[ExcelOrderRow],
) -> list[ReportMetric]:
    report_id = request.report_id
    filters = request.filters
    requested_source = _normalize_key(filters.source)

    if requested_source is not None and report_id is not ReportType.SALES_BY_SOURCE:
        raise ValueError(f"Source filter is not supported for report_id={report_id.value}")

    if report_id is ReportType.SALES_TOTAL:
        return [ReportMetric(label="sales_total", value=float(_sales_total(rows)))]

    if report_id is ReportType.ORDER_COUNT:
        return [ReportMetric(label="order_count", value=float(_order_count(rows)))]

    if report_id is ReportType.AVERAGE_CHECK:
        order_count = _order_count(rows)
        if order_count == 0:
            average = Decimal("0")
        else:
            average = _sales_total(rows) / Decimal(order_count)
        return [ReportMetric(label="average_check", value=float(average))]

    if report_id is ReportType.SALES_BY_SOURCE:
        source_totals = _group_sales(rows, key_getter=lambda row: row.source)
        if requested_source is not None:
            normalized_totals = {
                _normalize_key(label) or "unknown": value
                for label, value in source_totals.items()
            }
            normalized_catalog = {
                _normalize_key(row.source) or "unknown"
                for row in all_rows
            }
            if requested_source in normalized_totals:
                return [
                    ReportMetric(
                        label=requested_source,
                        value=float(normalized_totals[requested_source]),
                    )
                ]
            if requested_source not in normalized_catalog:
                raise ValueError(f"Unsupported source: {requested_source}")
            return [ReportMetric(label=requested_source, value=0.0)]
        metrics = _ranked_metrics(source_totals)
        return metrics or [ReportMetric(label="unknown", value=0.0)]

    if report_id is ReportType.SALES_BY_COURIER:
        metrics = _ranked_metrics(_group_sales(rows, key_getter=lambda row: row.courier))
        return metrics or [ReportMetric(label="unknown", value=0.0)]

    if report_id is ReportType.TOP_LOCATIONS:
        metrics = _ranked_metrics(
            _group_sales(rows, key_getter=lambda row: row.address),
            top_n=10,
        )
        return metrics or [ReportMetric(label="unknown", value=0.0)]

    if report_id is ReportType.TOP_CUSTOMERS:
        metrics = _ranked_metrics(_group_sales_by_customer(rows), top_n=10)
        return metrics or [ReportMetric(label="unknown", value=0.0)]

    if report_id is ReportType.REPEAT_CUSTOMER_RATE:
        customer_counts: Counter[str] = Counter(
            customer_key
            for customer_key in (_customer_identity_key(row) for row in rows)
            if customer_key is not None
        )
        distinct_customers = len(customer_counts)
        repeat_customers = sum(1 for count in customer_counts.values() if count > 1)
        rate = (
            Decimal(repeat_customers) * Decimal("100") / Decimal(distinct_customers)
            if distinct_customers > 0
            else Decimal("0")
        )
        return [
            ReportMetric(label="repeat_customer_rate_percent", value=float(rate)),
            ReportMetric(label="repeat_customer_count", value=float(repeat_customers)),
            ReportMetric(label="distinct_customer_count", value=float(distinct_customers)),
        ]

    if report_id is ReportType.DELIVERY_FEE_ANALYTICS:
        delivery_fees = [row.delivery_fee for row in rows if row.delivery_fee is not None]
        total_fee = sum((fee for fee in delivery_fees), start=Decimal("0"))
        avg_fee = total_fee / Decimal(len(delivery_fees)) if delivery_fees else Decimal("0")
        charged_count = sum(1 for fee in delivery_fees if fee > 0)
        return [
            ReportMetric(label="delivery_fee_total", value=float(total_fee)),
            ReportMetric(label="delivery_fee_average", value=float(avg_fee)),
            ReportMetric(label="delivery_fee_charged_orders", value=float(charged_count)),
        ]

    if report_id is ReportType.PAYMENT_COLLECTION:
        invoiced_total = sum(
            (row.invoice for row in rows if row.invoice is not None),
            start=Decimal("0"),
        )
        paid_total = sum(
            (row.paid_amount for row in rows if row.paid_amount is not None),
            start=Decimal("0"),
        )
        collection_rate = (
            paid_total * Decimal("100") / invoiced_total
            if invoiced_total != 0
            else Decimal("0")
        )
        return [
            ReportMetric(label="invoiced_total", value=float(invoiced_total)),
            ReportMetric(label="paid_total", value=float(paid_total)),
            ReportMetric(label="collection_rate_percent", value=float(collection_rate)),
        ]

    if report_id is ReportType.OUTSTANDING_BALANCE:
        invoiced_total = sum(
            (row.invoice for row in rows if row.invoice is not None),
            start=Decimal("0"),
        )
        paid_total = sum(
            (row.paid_amount for row in rows if row.paid_amount is not None),
            start=Decimal("0"),
        )
        return [
            ReportMetric(label="outstanding_balance", value=float(invoiced_total - paid_total)),
        ]

    if report_id is ReportType.DAILY_SALES_TREND:
        metrics = _daily_sales_metrics(rows)
        return metrics or [ReportMetric(label=filters.date_from.isoformat(), value=0.0)]

    if report_id is ReportType.DAILY_ORDER_TREND:
        metrics = _daily_order_metrics(rows)
        return metrics or [ReportMetric(label=filters.date_from.isoformat(), value=0.0)]

    if report_id is ReportType.SALES_BY_WEEKDAY:
        return _sales_by_weekday_metrics(rows)

    if report_id is ReportType.GROSS_PROFIT:
        sales_total = _sales_total(rows)
        net_cost_total = sum(
            (row.net_cost for row in rows if row.net_cost is not None),
            start=Decimal("0"),
        )
        gross_profit = sales_total - net_cost_total
        gross_margin = (
            gross_profit * Decimal("100") / sales_total
            if sales_total != 0
            else Decimal("0")
        )
        return [
            ReportMetric(label="gross_profit", value=float(gross_profit)),
            ReportMetric(label="gross_margin_percent", value=float(gross_margin)),
        ]

    if report_id is ReportType.LOCATION_CONCENTRATION:
        location_totals = _group_sales(rows, key_getter=lambda row: row.address)
        total_sales = sum(location_totals.values(), start=Decimal("0"))
        ranked = sorted(location_totals.values(), reverse=True)
        top_10_sales = sum(ranked[:10], start=Decimal("0"))
        top_1_sales = ranked[0] if ranked else Decimal("0")
        top_10_share = (
            top_10_sales * Decimal("100") / total_sales
            if total_sales != 0
            else Decimal("0")
        )
        top_1_share = (
            top_1_sales * Decimal("100") / total_sales
            if total_sales != 0
            else Decimal("0")
        )
        return [
            ReportMetric(label="top_10_location_share_percent", value=float(top_10_share)),
            ReportMetric(label="top_1_location_share_percent", value=float(top_1_share)),
            ReportMetric(label="distinct_locations_count", value=float(len(location_totals))),
        ]

    raise ValueError(f"Unsupported report_id: {report_id.value}")


def run_excel_report(
    request: ReportRequest,
    *,
    file_path: str | Path,
    sheet_name: str | None = None,
) -> RunReportResponse:
    rows = load_excel_orders(file_path=file_path, sheet_name=sheet_name)
    filtered_rows = _rows_in_date_range(rows, request.filters)
    metrics = _metrics_for_request(request, filtered_rows, all_rows=rows)
    result = ReportResult(
        report_id=request.report_id,
        filters=request.filters,
        metrics=metrics,
        generated_at=_generated_at(request.filters),
    )
    return RunReportResponse(result=result, warnings=[EXCEL_BACKEND_WARNING])
